import torch
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.patches as patches
from PIL import Image, ImageDraw
from matplotlib.patches import Patch
from transformers import AutoModelForObjectDetection, TableTransformerForObjectDetection
from torchvision import transforms
from vietocr.tool.predictor import Predictor
from vietocr.tool.config import Cfg
from tqdm.auto import tqdm
from openpyxl import load_workbook

# Định nghĩa lớp MaxResize để thay đổi kích thước hình ảnh
class MaxResize(object):
    def __init__(self, max_size=800):
        self.max_size = max_size

    def __call__(self, image):
        width, height = image.size
        current_max_size = max(width, height)
        scale = self.max_size / current_max_size
        resized_image = image.resize((int(round(scale*width)), int(round(scale*height))))
        return resized_image

# Hàm điều chỉnh bounding box
def adjust_bbox(bbox, img_size, margin=0.02):
    x1, y1, x2, y2 = bbox
    img_w, img_h = img_size
    width = x2 - x1
    height = y2 - y1
    dx = width * margin
    dy = height * margin
    x1 = max(0, x1 - dx)
    y1 = max(0, y1 - dy)
    x2 = min(img_w, x2 + dx)
    y2 = min(img_h, y2 + dy)
    return [x1, y1, x2, y2]

# Hàm chuyển đổi box từ định dạng (center, width, height) sang (x1, y1, x2, y2)
def box_cxcywh_to_xyxy(x):
    x_c, y_c, w, h = x.unbind(-1)
    b = [(x_c - 0.5 * w), (y_c - 0.5 * h), (x_c + 0.5 * w), (y_c + 0.5 * h)]
    return torch.stack(b, dim=1)

# Hàm thay đổi kích thước bounding box theo kích thước ảnh
def rescale_bboxes(out_bbox, size):
    img_w, img_h = size
    b = box_cxcywh_to_xyxy(out_bbox)
    b = b * torch.tensor([img_w, img_h, img_w, img_h], dtype=torch.float32)
    return b

# Hàm chuyển đổi kết quả đầu ra thành các đối tượng (bảng)
def outputs_to_objects(outputs, img_size, id2label):
    m = outputs.logits.softmax(-1).max(-1)
    pred_labels = list(m.indices.detach().cpu().numpy())[0]
    pred_scores = list(m.values.detach().cpu().numpy())[0]
    pred_bboxes = outputs['pred_boxes'].detach().cpu()[0]
    pred_bboxes = [elem.tolist() for elem in rescale_bboxes(pred_bboxes, img_size)]
    
    objects = []
    for label, score, bbox in zip(pred_labels, pred_scores, pred_bboxes):
        class_label = id2label[int(label)]
        if not class_label == 'no object':
            objects.append({'label': class_label, 'score': float(score),
                            'bbox': [float(elem) for elem in bbox]})
            
    return objects

# Hàm chuyển đổi một hình ảnh Matplotlib thành hình ảnh PIL
def fig2img(fig):
    import io
    buf = io.BytesIO()
    fig.savefig(buf, format='png')
    buf.seek(0)
    img = Image.open(buf)
    return img

# Hàm hiển thị các bảng đã được phát hiện
def visualize_detected_tables(img, det_tables, out_path=None):
    
    plt.imshow(img, interpolation="lanczos")
    fig = plt.gcf()
    fig.set_size_inches(20, 20)
    ax = plt.gca()
    img_size = img.size

    for det_table in det_tables:
        bbox = det_table['bbox']
        bbox = adjust_bbox(bbox, img_size)
        if det_table['label'] == 'table':
            facecolor = (1, 0, 0.45)
            edgecolor = (1, 0, 0.45)
            alpha = 0.3
            linewidth = 2
            hatch='//////'
        elif det_table['label'] == 'table rotated':
            facecolor = (0.95, 0.6, 0.1)
            edgecolor = (0.95, 0.6, 0.1)
            alpha = 0.3
            linewidth = 2
            hatch='//////'
        else:
            continue
        
        rect = patches.Rectangle(bbox[:2], bbox[2]-bbox[0], bbox[3]-bbox[1], linewidth=linewidth,
                                    edgecolor='none',facecolor=facecolor, alpha=0.1)
        ax.add_patch(rect)
        rect = patches.Rectangle(bbox[:2], bbox[2]-bbox[0], bbox[3]-bbox[1], linewidth=linewidth,
                                    edgecolor=edgecolor,facecolor='none',linestyle='-', alpha=alpha)
        ax.add_patch(rect)
        rect = patches.Rectangle(bbox[:2], bbox[2]-bbox[0], bbox[3]-bbox[1], linewidth=0,
                                    edgecolor=edgecolor,facecolor='none',linestyle='-', hatch=hatch, alpha=0.2)
        ax.add_patch(rect)
        
    plt.xticks([], [])
    plt.yticks([], [])
    
    legend_elements = [Patch(facecolor=(1, 0, 0.45), edgecolor=(1, 0, 0.45),
                                label='Table', hatch='//////', alpha=0.3),
                        Patch(facecolor=(0.95, 0.6, 0.1), edgecolor=(0.95, 0.6, 0.1),
                                label='Table (rotated)', hatch='//////', alpha=0.3)]
    plt.legend(handles=legend_elements, bbox_to_anchor=(0.5, -0.02), loc='upper center', borderaxespad=0,
                    fontsize=10, ncol=2)
    plt.gcf().set_size_inches(7, 5)
    plt.axis('off')
    if out_path is not None:
      plt.savefig(out_path, bbox_inches='tight', dpi=120)
    return fig

# Hàm cắt các bảng từ ảnh
def objects_to_crops(img, tokens, objects, class_thresholds, padding=10):
    table_crops = []
    img_w, img_h = img.size
    
    for obj in objects:
        if obj['score'] < class_thresholds[obj['label']]:
            continue
        cropped_table = {}
        bbox = obj['bbox']
        bbox = [bbox[0] - padding, bbox[1] - padding, bbox[2] + padding, bbox[3] + padding]
        bbox = [
            max(0, bbox[0]),
            max(0, bbox[1]),
            min(img_w, bbox[2]),
            min(img_h, bbox[3])
        ]
        cropped_img = img.crop(bbox)
        table_tokens = [token for token in tokens if iob(token['bbox'], bbox) >= 0.5]
        for token in table_tokens:
            token['bbox'] = [
                token['bbox'][0]-bbox[0],
                token['bbox'][1]-bbox[1],
                token['bbox'][2]-bbox[0],
                token['bbox'][3]-bbox[1]]
        if obj['label'] == 'table rotated':
            cropped_img = cropped_img.rotate(270, expand=True)
            for token in table_tokens:
                bbox = token['bbox']
                bbox = [cropped_img.size[0]-bbox[3]-1,
                        bbox[0],
                        cropped_img.size[0]-bbox[1]-1,
                        bbox[2]]
                token['bbox'] = bbox
                
        cropped_table['image'] = cropped_img
        cropped_table['tokens'] = table_tokens
        table_crops.append(cropped_table)
        
    return table_crops

# Hàm lấy tọa độ các ô dựa vào hàng
def get_cell_coordinates_by_row(table_data):
    rows = [entry for entry in table_data if entry['label'] == 'table row']
    columns = [entry for entry in table_data if entry['label'] == 'table column']
    rows.sort(key=lambda x: x['bbox'][1])
    columns.sort(key=lambda x: x['bbox'][0])
    
    def find_cell_coordinates(row, column,padding=4):
        ##cell_bbox = [column['bbox'][0], row['bbox'][1], column['bbox'][2], row['bbox'][3]]
        
        x1, y1 = column['bbox'][0], row['bbox'][1]
        x2, y2 = column['bbox'][2], row['bbox'][3]

        # Áp dụng padding
        x1 -= padding -2 
        y1 -= padding
        
        x2 += 2
        y2 += padding + 2
        
        cell_bbox = [x1, y1, x2, y2]
        return cell_bbox
    
    cell_coordinates = []
    
    for row in rows:
        row_cells = []
        for column in columns:
            cell_bbox = find_cell_coordinates(row, column)
            row_cells.append({'column': column['bbox'], 'cell': cell_bbox})
        row_cells.sort(key=lambda x: x['column'][0])
        cell_coordinates.append({'row': row['bbox'], 'cells': row_cells, 'cell_count': len(row_cells)})
    cell_coordinates.sort(key=lambda x: x['row'][1])
    
    return cell_coordinates

# Hàm áp dụng OCR cho các ô
def apply_ocr(cell_coordinates, cropped_table, predictor):
    data = dict()
    max_num_columns = 0
    
    for idx, row in enumerate(tqdm(cell_coordinates)):
        row_text = []
        for cell in row["cells"]:
        
            #Crop cell from image
            cell_image = cropped_table.crop(cell["cell"])
            cell_image = np.array(cell_image)
            cell_image_pil = Image.fromarray(cell_image)
            # plt.imshow(cell_image_pil)  # turn on to more understand
            # plt.show()
            result = predictor.predict(cell_image_pil)
            print(" ", result)
            
    
            if len(result) > 0: #if result:
                text = "".join(result)
                #print(text)
                row_text.append(text)
                
                
        if len(row_text) > max_num_columns:
            max_num_columns = len(row_text)
            
        data[idx] = row_text
        
    for row, row_data in data.copy().items():
        if len(row_data) != max_num_columns:
            row_data = row_data + ["" for _ in range(max_num_columns - len(row_data))]
        data[row] = row_data
        
    return data

# Hàm hiển thị các hàng trong bảng
def plot_results(cells,structure_model,cropped_table, class_to_visualize):
    if class_to_visualize not in structure_model.config.id2label.values():
        raise ValueError("Class should be one of the available classes")
    
    plt.figure(figsize=(8,4))
    plt.imshow(cropped_table)
    ax = plt.gca()
    
    for cell in cells:
        score = cell["score"]
        bbox = cell["bbox"]
        label = cell["label"]
        if label == class_to_visualize:
            xmin, ymin, xmax, ymax = tuple(bbox)
            ax.add_patch(plt.Rectangle((xmin, ymin), xmax - xmin, ymax - ymin, fill=False, color="red", linewidth=2))
            text = f'{cell["label"]}: {score:0.2f}'
            ax.text(xmin, ymin, text, fontsize=12, bbox=dict(facecolor='yellow', alpha=0.5)) #alpha độ trong suốt
            
    plt.axis('off')
    plt.show()
import pandas as pd
from openpyxl import load_workbook

def save_ocr_results_to_excel(data, output_path='output/output_vietocr.xlsx'):
    """
    Lưu kết quả OCR thành file Excel và điều chỉnh độ rộng cột tự động.

    Parameters:
    - data: Dữ liệu kết quả OCR dưới dạng dictionary.
    - output_path: Đường dẫn lưu file Excel.
    """
    # Tạo DataFrame từ dữ liệu OCR
    df = pd.DataFrame.from_dict(data, orient='index')
    
    # Lưu DataFrame thành file Excel
    df.to_excel(output_path, index=False)
    
    # Mở lại file Excel vừa tạo để điều chỉnh độ rộng cột
    workbook = load_workbook(output_path)
    worksheet = workbook.active
    
    # Tự động điều chỉnh độ rộng các cột
    for column in worksheet.columns:
        max_length = 0
        column = list(column)
        for cell in column:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = max_length + 2
        worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
    
    # Lưu lại file Excel sau khi điều chỉnh độ rộng cột
    workbook.save(output_path)

